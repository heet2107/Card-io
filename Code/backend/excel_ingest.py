"""
CardioReport – Excel Ingestion
Robust loader that handles the PAMHealth study files.
Patient discovery is delegated to data_registry_v2 (filename-first parsing).
Multi-sensor patients (EG, JB, RSanchez) are merged under a shared key.
"""

from __future__ import annotations
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
import pandas as pd

from .config import settings, RR_NOISE_THRESHOLD_WHEN_HR_MISSING
from .data_registry_v2 import scan_directory, PATIENT_GROUPS


def apply_rr_noise_filter(df: pd.DataFrame) -> pd.DataFrame:
    """R22.A — zero RR samples that are sensor noise rather than real breathing.

    When the radar loses HR lock it can emit spurious RR values (motion in
    the room, signal dropout). Real high RR co-occurs with a valid HR, so a
    high RR with HR missing/zero is treated as noise and zeroed out before
    episode detection runs.
    """
    if df.empty or "rr_avg" not in df.columns or "hr_avg" not in df.columns:
        return df
    hr_missing = df["hr_avg"].isna() | (df["hr_avg"] == 0)
    rr_high = df["rr_avg"].fillna(0) > RR_NOISE_THRESHOLD_WHEN_HR_MISSING
    mask = hr_missing & rr_high
    rr_cols = [c for c in ("rr_avg", "rr_min", "rr_max") if c in df.columns]
    if mask.any() and rr_cols:
        df.loc[mask, rr_cols] = 0
    return df


# ── Registry Cache ───────────────────────────────────────────────────────────

_registry_cache: dict | None = None


def _get_registry(excel_path: Optional[str] = None) -> dict:
    """Return (and cache) the patient registry built by scan_directory."""
    global _registry_cache
    target = str(Path(excel_path or settings.excel_path).resolve())
    if _registry_cache is None:
        _registry_cache = scan_directory(target)
    return _registry_cache


def _device_to_patient_id(device_id: str, patient_name: str) -> str:
    """Map a raw device_id to the canonical patient key used in the registry."""
    for group_id, group in PATIENT_GROUPS.items():
        if device_id in group['devices']:
            return group_id
    return f'{device_id}_{patient_name}'



# ── Column Mapping ───────────────────────────────────────────────────────────

# The real Excel uses these header names (with leading spaces):
_COL_MAP_REAL = {
    "Date": "date",
    " Time": "time",
    " XL-time": "timestamp",
    " avg_Hr": "hr_avg",
    "  max_Hr": "hr_max",
    " min_Hr": "hr_min",
    "avg_Rr": "rr_avg",
    " max_Rr": "rr_max",
    " min_Rr": "rr_min",
    "cnt": "cnt",
}

# Normalised names (for format A with clean headers)
_COL_MAP_CLEAN = {
    "timestamp": "timestamp",
    "hr_avg": "hr_avg",
    "hr_max": "hr_max",
    "hr_min": "hr_min",
    "rr_avg": "rr_avg",
    "rr_max": "rr_max",
    "rr_min": "rr_min",
    "cnt": "cnt",
    "gap_flag": "gap_flag",
    "patient_id": "patient_id",
    # Aliases
    "Hr": "hr_avg",
    "Rr": "rr_avg",
    "HR": "hr_avg",
    "RR": "rr_avg",
    "Date-Time": "timestamp",
}


def _extract_patient_id_from_sheet(name: str) -> str:
    """Extract patient ID from the sheet name (e.g. '934297-0122 J.LivRm_PVitalsag_2')."""
    # Pattern: 123456-1234
    match = re.match(r"(\d{4,}-\d+)", name)
    if match:
        return match.group(1)
    
    # Pattern: Juanita-Bed-10202023 -> Juanita-Bed
    parts = name.split("-")
    if len(parts) >= 2:
        # If last part is date-like (8 digits), exclude it
        if len(parts[-1]) == 8 and parts[-1].isdigit():
            return "-".join(parts[:-1])
        return "-".join(parts[:2])
    
    return name.strip()[:30]


def _normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map raw Excel columns to the canonical internal names."""
    rename = {}
    cols = list(df.columns)
    for raw, canon in _COL_MAP_REAL.items():
        if raw in cols:
            rename[raw] = canon
    for raw, canon in _COL_MAP_CLEAN.items():
        if raw in cols and raw not in rename:
            rename[raw] = canon
    # case-insensitive fallback
    lower_map = {c.strip().lower(): c for c in cols}
    for canon in ["hr_avg", "hr_max", "hr_min", "rr_avg", "rr_max", "rr_min", "cnt", "timestamp"]:
        if canon not in rename.values():
            if canon in lower_map:
                rename[lower_map[canon]] = canon
    df = df.rename(columns=rename)
    
    # Ensure no duplicate canonical columns (if multiple raw cols mapped to same canon)
    # We keep the first one
    df = df.loc[:, ~df.columns.duplicated()]
    
    return df


def _build_timestamp(df: pd.DataFrame) -> pd.DataFrame:
    """Build a proper datetime 'timestamp' column from available date/time cols."""
    if "timestamp" in df.columns and pd.api.types.is_datetime64_any_dtype(df["timestamp"]):
        return df

    # If timestamp column exists but isn't datetime yet, try to convert it directly first
    if "timestamp" in df.columns:
        converted = pd.to_datetime(df["timestamp"], errors="coerce")
        if converted.notna().sum() > len(df) * 0.5:
            df["timestamp"] = converted
            return df

    # Fallback: combine date + time columns
    if "date" in df.columns and "time" in df.columns:
        def _combine(row):
            d = row["date"]
            t = row["time"]
            if isinstance(d, datetime):
                d = d.date()
            if pd.isna(d) or pd.isna(t):
                return pd.NaT
            from datetime import time as _time
            if isinstance(t, _time):
                return datetime.combine(d, t)
            if isinstance(t, datetime):
                return datetime.combine(d, t.time())
            # Handle string time like ' 06:00:00'
            if isinstance(t, str):
                t = t.strip()
                try:
                    parsed_t = datetime.strptime(t, "%H:%M:%S").time()
                    return datetime.combine(d, parsed_t)
                except ValueError:
                    pass
            return pd.NaT

        df["timestamp"] = df.apply(_combine, axis=1)

    elif "timestamp" in df.columns:
        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")

    return df


def _extract_location_from_sheet(name: str) -> str:
    """Extract location from sheet name (e.g. LivRm, Chair, Bed)."""
    from .config import Locations
    name_lower = name.lower()
    if any(x in name_lower for x in ["livrm", "living room", "living_room"]):
        return Locations.LIVING_ROOM
    if "chair" in name_lower:
        return Locations.CHAIR
    if "bed" in name_lower:
        return Locations.BED
    return Locations.UNKNOWN


# ── Main Loader ──────────────────────────────────────────────────────────────

_cache: dict[str, pd.DataFrame] = {}


def load_vitals(excel_path: Optional[str] = None, force_reload: bool = False) -> dict[str, pd.DataFrame]:
    """
    Load vitals from Excel(s).
    If excel_path is a directory, loads all .xlsx files in it.
    Returns { patient_id: DataFrame } where each DF has normalised columns:
        patient_id, timestamp, hr_avg, hr_max, hr_min, rr_avg, rr_max, rr_min, cnt, gap_flag, location
    Sorted by timestamp.
    """
    global _cache
    target = Path(excel_path or settings.excel_path)
    cache_key = str(target)

    if not force_reload and cache_key in _cache:
        full = _cache[cache_key]
        if "patient_id" in full.columns:
            return {pid: grp.reset_index(drop=True) for pid, grp in full.groupby("patient_id")}
        return {"unknown": full}

    # Decide which files to load
    files_to_load = []
    if target.is_dir():
        files_to_load = list(target.glob("*.xlsx"))
    elif target.is_file():
        files_to_load = [target]
    else:
        # Fallback to the directory of the file if it doesn't exist but its parent does
        if target.parent.exists():
            files_to_load = list(target.parent.glob("*.xlsx"))

    if not files_to_load:
        raise FileNotFoundError(f"No Excel files found at {target}")

    all_frames: list[pd.DataFrame] = []

    for path in files_to_load:
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                # Skip non-aggregate sheets (Summary, Low HR Events, per-day raw data)
                sn_lower = sheet_name.strip().lower()
                if sn_lower in ("summary", "low hr events"):
                    continue
                # Skip date-named raw-data sheets (e.g. Juanita-Bed-10202023, Juanita-Bed-11xx2023)
                if re.search(r'[\dx]{6,}$', sheet_name.replace(" ", "")):
                    continue

                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue
                header = [str(c) if c is not None else f"_col_{i}" for i, c in enumerate(rows[0])]
                data = rows[1:]

                if data and all(v is None or (isinstance(v, str) and v.strip() == "") for v in data[0]):
                    data = data[1:]

                # Find useful columns: stop at 2 consecutive unnamed columns after index 5
                useful = len(header)
                _known_cols = {"date", "time", "xl-time", "avg_hr", "max_hr", "min_hr",
                               "avg_rr", "max_rr", "min_rr", "cnt"}
                consecutive_unknown = 0
                for i, h in enumerate(header):
                    h_clean = h.strip().lower().replace(" ", "").replace("_", "")
                    is_known = any(k.replace("_", "") in h_clean for k in _known_cols) or h.strip() == ""
                    if h.startswith("_col_") and i > 5:
                        consecutive_unknown += 1
                        if consecutive_unknown >= 2:
                            useful = i - 1
                            break
                    else:
                        consecutive_unknown = 0
                
                # Deduplicate columns if any (happens in Juanita file)
                seen_cols = {}
                final_header = []
                for h in header[:useful]:
                    if h in seen_cols:
                        seen_cols[h] += 1
                        final_header.append(f"{h}_{seen_cols[h]}")
                    else:
                        seen_cols[h] = 0
                        final_header.append(h)
                header = final_header
                data = [row[:useful] for row in data]

                df = pd.DataFrame(data, columns=header)
                df = df.dropna(how="all")
                if df.empty: continue
                
                first_col = header[0]
                df = df[df[first_col].notna()]

                df = _normalise_columns(df)
                df = _build_timestamp(df)

                for col in ["hr_avg", "hr_max", "hr_min", "rr_avg", "rr_max", "rr_min", "cnt"]:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce")

                df = df.dropna(subset=["timestamp"])
                df = df[df["hr_avg"].notna() | df["rr_avg"].notna()]

                if "gap_flag" not in df.columns:
                    df["gap_flag"] = 0

                # ── Patient ID assignment (registry-first) ──────────────
                # Try to look up device_id from the registry for this file.
                # This assigns the friendly patient name and groups
                # multi-sensor patients (EG, JB, RSanchez) under one key.
                from .config import Locations
                registry = _get_registry(str(target))
                # Build display names so we can use the disambiguated name as PID
                display_name_map = _patient_display_names(str(target))
                assigned_pid: Optional[str] = None
                assigned_location: Optional[str] = None

                for rec_pid, rec in registry.items():
                    for sensor in rec.sensors:
                        if Path(sensor.file_path).resolve() == path.resolve():
                            # Use display name (handles disambiguation like "S (Chair)")
                            assigned_pid = display_name_map.get(rec_pid, rec.patient_name)
                            loc_map = {'bed': Locations.BED, 'chair': Locations.CHAIR}
                            assigned_location = loc_map.get(sensor.sensor_type, Locations.UNKNOWN)
                            break
                    if assigned_pid:
                        break

                if assigned_pid is None:
                    # Fallback: derive from sheet name (old behaviour)
                    assigned_pid = _extract_patient_id_from_sheet(sheet_name)

                if "patient_id" not in df.columns:
                    df["patient_id"] = assigned_pid

                # Location extraction
                if assigned_location:
                    df["location"] = assigned_location
                else:
                    df["location"] = _extract_location_from_sheet(sheet_name)
                    if df["location"].iloc[0] == Locations.UNKNOWN:
                        df["location"] = _extract_location_from_sheet(path.name)

                keep = ["patient_id", "timestamp", "hr_avg", "hr_max", "hr_min",
                        "rr_avg", "rr_max", "rr_min", "cnt", "gap_flag", "location"]
                keep = [c for c in keep if c in df.columns]
                df = df[keep].copy()

                all_frames.append(df)
            wb.close()
        except Exception as e:
            print(f"Warning: Failed to load {path}: {e}")

    if not all_frames:
        raise ValueError(f"No valid vitals data found in {target}")

    combined = pd.concat(all_frames, ignore_index=True)
    combined = combined.sort_values(["patient_id", "timestamp", "location"]).reset_index(drop=True)

    # Deduplicate: keep first occurrence of (patient_id, timestamp, location)
    before = len(combined)
    combined = combined.drop_duplicates(subset=["patient_id", "timestamp", "location"], keep="first")
    after = len(combined)
    if before != after:
        print(f"Deduplicated: {before} -> {after} rows ({before - after} duplicates removed)")

    combined = apply_rr_noise_filter(combined)

    _cache[cache_key] = combined

    if "patient_id" in combined.columns:
        return {pid: grp.reset_index(drop=True) for pid, grp in combined.groupby("patient_id")}
    return {"unknown": combined}


def _patient_display_names(excel_path: Optional[str] = None) -> dict:
    """Return a mapping of registry patient_id → display name.

    When two records share the same patient_name (e.g. both "S"), the
    sensor type is appended: "S (Chair)" / "S (Bed)".
    """
    registry = _get_registry(excel_path)
    # Count how many records share each raw patient_name
    from collections import Counter
    name_counts: Counter = Counter(rec.patient_name for rec in registry.values())

    result = {}
    for pid, rec in registry.items():
        if name_counts[rec.patient_name] > 1:
            # Disambiguate: append sensor type(s)
            types_str = ' + '.join(t.capitalize() for t in rec.sensor_types)
            result[pid] = f"{rec.patient_name} ({types_str})"
        else:
            result[pid] = rec.patient_name
    return result


def get_patient_ids(excel_path: Optional[str] = None) -> list[str]:
    """Return sorted list of *display* patient names from the registry.

    Registry-based discovery means names come from filenames (PHolst,
    Wimberley, EG …) instead of raw device IDs.  Patients who share a
    name (e.g. two "S" patients) are disambiguated with a sensor suffix.
    """
    display_names = _patient_display_names(excel_path)
    return sorted(display_names.values())


def get_patient_metadata(patient_id: str, excel_path: Optional[str] = None) -> dict:
    """Return location/sensor and date range metadata for a patient.

    Uses the registry for sensor_types and date ranges (fast, no full load);
    only falls back to the loaded DataFrame for location strings.

    Returns:
        {
            "patient_id": str,
            "locations": ["Chair", "Bed", ...],
            "sensor_types": ["chair"] | ["bed"] | ["bed","chair"],
            "date_range": {"start": "2024-05-01", "end": "2024-06-30"},
            "total_hours": int,
        }
    """
    from .config import Locations

    # Try to find the record in the registry by matching display name, patient_name, or patient_id
    registry = _get_registry(excel_path)
    display_name_map = _patient_display_names(excel_path)  # {registry_pid: display_name}
    reg_rec = None
    for rec_pid, rec in registry.items():
        display_name = display_name_map.get(rec_pid, rec.patient_name)
        if display_name == patient_id or rec.patient_name == patient_id or rec.patient_id == patient_id:
            reg_rec = rec
            break

    if reg_rec is not None:
        # Build location list from registry sensor types
        SENSOR_TO_LOC = {'bed': Locations.BED, 'chair': Locations.CHAIR}
        locations = sorted(
            {SENSOR_TO_LOC[st] for st in reg_rec.sensor_types if st in SENSOR_TO_LOC}
        )
        dr = reg_rec.date_range
        date_range = {
            "start": dr[0].strftime("%Y-%m-%d") if dr[0] else "",
            "end":   dr[1].strftime("%Y-%m-%d") if dr[1] else "",
        }
        total_hours = sum(s.total_hours for s in reg_rec.sensors)
        return {
            "patient_id": patient_id,
            "locations": locations,
            "sensor_types": reg_rec.sensor_types,
            "date_range": date_range,
            "total_hours": total_hours,
        }

    # Fallback: derive from loaded DataFrame (old behaviour)
    data = load_vitals(excel_path)
    if patient_id not in data:
        return {"patient_id": patient_id, "locations": [], "sensor_types": [], "date_range": {}, "total_hours": 0}

    df = data[patient_id]
    locations = []
    if "location" in df.columns:
        raw_locs = sorted(df["location"].unique())
        locations = [loc for loc in raw_locs if loc != Locations.UNKNOWN]

    ts = df["timestamp"]
    date_range = {
        "start": ts.min().strftime("%Y-%m-%d") if len(ts) > 0 else "",
        "end":   ts.max().strftime("%Y-%m-%d") if len(ts) > 0 else "",
    }
    return {
        "patient_id": patient_id,
        "locations": locations,
        "sensor_types": [],
        "date_range": date_range,
        "total_hours": len(df),
    }


# ── Bed Sensor Data Loaders ─────────────────────────────────────────────────

def _find_bed_excel(excel_path: Optional[str] = None) -> Optional[Path]:
    """Find Excel file(s) that contain the 'Summary' and 'Low HR Events' sheets (bed sensor)."""
    target = Path(excel_path or settings.excel_path)
    files = list(target.glob("*.xlsx")) if target.is_dir() else [target]
    for f in files:
        try:
            wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
            sheets_lower = [s.lower() for s in wb.sheetnames]
            wb.close()
            if "summary" in sheets_lower and "low hr events" in sheets_lower:
                return f
        except Exception:
            continue
    return None


def load_bed_summary(excel_path: Optional[str] = None) -> Optional[pd.DataFrame]:
    """Load daily bed-time summary from the 'Summary' sheet of the bed sensor Excel.

    Returns DataFrame with columns:
        date, hr_avg, hr_low, bedtime_start, bedtime_end, hours_in_bed
    Or None if no bed sensor file found.
    """
    bed_file = _find_bed_excel(excel_path)
    if bed_file is None:
        return None

    wb = openpyxl.load_workbook(str(bed_file), read_only=True, data_only=True)
    ws = wb["Summary"]

    rows_data = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows_data) < 4:
        return None

    # Headers are in row index 2 (0-indexed)
    # Data starts at row index 3
    records = []
    for row in rows_data[3:]:
        vals = list(row)
        date_val = vals[1]
        hr_avg = vals[2]
        hr_low = vals[4]  # HR_low is column index 4
        bedtime_start = vals[16]
        bedtime_end = vals[17]
        hours_in_bed = vals[18]

        # Skip non-date rows (e.g. "Average" summary row)
        if date_val is None or isinstance(date_val, str):
            continue
        if not isinstance(date_val, datetime):
            continue

        records.append({
            "date": pd.Timestamp(date_val).normalize(),
            "hr_avg": float(hr_avg) if hr_avg is not None else np.nan,
            "hr_low": float(hr_low) if hr_low is not None else np.nan,
            "bedtime_start": bedtime_start,
            "bedtime_end": bedtime_end,
            "hours_in_bed": float(hours_in_bed) if hours_in_bed is not None else np.nan,
        })

    if not records:
        return None

    return pd.DataFrame(records)


def load_low_hr_alerts(excel_path: Optional[str] = None) -> Optional[pd.DataFrame]:
    """Load low HR alert events from the 'Low HR Events' sheet.

    Returns DataFrame with columns:
        timestamp, alert_text, alert_hr
    Or None if no bed sensor file found.
    """
    bed_file = _find_bed_excel(excel_path)
    if bed_file is None:
        return None

    wb = openpyxl.load_workbook(str(bed_file), read_only=True, data_only=True)
    ws = wb["Low HR Events"]

    rows_data = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows_data) < 2:
        return None

    records = []
    for row in rows_data[1:]:
        vals = list(row)
        timestamp_str = vals[0]  # ISO timestamp string
        alert_text = vals[6]     # Atext column
        alert_hr = vals[12]      # Alert_HR column

        if timestamp_str is None:
            continue

        # Parse ISO timestamp
        try:
            if isinstance(timestamp_str, str):
                ts = pd.Timestamp(timestamp_str)
            elif isinstance(timestamp_str, datetime):
                ts = pd.Timestamp(timestamp_str)
            else:
                continue
        except Exception:
            continue

        records.append({
            "timestamp": ts,
            "alert_text": str(alert_text).strip() if alert_text else "Low HR",
            "alert_hr": int(alert_hr) if alert_hr is not None else 0,
        })

    if not records:
        return None

    return pd.DataFrame(records)


def has_bed_sensor(patient_id: str, excel_path: Optional[str] = None) -> bool:
    """Check if this patient has bed sensor data.

    Uses the registry for O(1) lookup; falls back to DataFrame scan.
    """
    from .config import Locations
    registry = _get_registry(excel_path)
    for rec in registry.values():
        if rec.patient_name == patient_id or rec.patient_id == patient_id:
            return rec.has_bed

    # Fallback
    data = load_vitals(excel_path)
    if patient_id not in data:
        return False
    df = data[patient_id]
    if "location" not in df.columns:
        return False
    locations = df["location"].unique()
    return Locations.BED in locations
