"""
CardioReport – FastAPI Main Application

Six-stage pipeline per Implementation Guide Section 9:
  Stage 1: Ingest (excel_ingest.py)
  Stage 2: Compute (signal_engine.py)
  Stage 3: Detect (episodes.py)
  Stage 4: Narrate (narrative_ai.py) — triage/trend computed BEFORE AI
  Stage 5: Chart (charts.py)
  Stage 6: Render (pdf_render.py)
"""

from __future__ import annotations
import hashlib
from datetime import datetime
from typing import Optional, Literal

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response

from cachetools import TTLCache

from .config import settings, Locations
from .models import (
    ReportRequest, ReportResponse,
    BedDaySummary, BedActivitySummary,
)
from .excel_ingest import (
    load_vitals, get_patient_ids, get_patient_metadata,
    load_bed_summary, load_low_hr_alerts, _find_bed_excel,
)
from .signal_engine import (
    apply_window, compute_stats, compute_full_stats, compute_data_quality,
    compute_data_resolution, compute_triage, compute_trend_assessment,
    compute_action_posture, compute_positional_stats, compute_activity_data,
    compute_end_of_period_clustering, compute_last_24h_snapshot,
)
from .triage_24h import compute_24h_layer
from .medhab_ingest import load_medhab_vitals, discover_report_windows
from .client_registry import (
    discover_clients, client_label, client_is_csv, client_specs,
    load_client_data, list_patients, resolve_client_for_patient,
    is_patient_in_client,
)
from .episodes import detect_episodes, compute_rollups
from .narrative_ai import generate_narrative
from .charts import (
    generate_combined_chart, generate_histogram,
    generate_positional_chart, generate_activity_trend_chart,
    generate_bed_hours_chart,
)
from .pdf_render import generate_pdf


# ── App setup ────────────────────────────────────────────────────────────────

app = FastAPI(
    title="CardioReport API",
    version=settings.app_version,
    description="RPM trend report engine",
)


@app.on_event("startup")
async def _warm_client_caches():
    """Round 28 — pre-warm each library's data cache on startup so the first
    patient-list load is instant. The Excel-shape PAM cohort is a ~7s cold
    registry scan; without warming, the first 'PAM Health' selection reads as a
    frozen dropdown. Runs off the event loop so startup isn't blocked, and is
    best-effort (a failing client must not stop the server). Awaited on a
    worker thread so startup completes only once caches are warm — the first
    real selection is then guaranteed instant, with no warm-vs-request race."""
    import asyncio

    def _warm():
        for c in discover_clients():
            try:
                load_client_data(c)
                client_specs(c)
                print(f"[warm] library '{c}' pre-loaded")
            except Exception as e:  # never let a bad folder block boot
                print(f"[warm] library '{c}' failed to pre-load: {e}")

    await asyncio.get_event_loop().run_in_executor(None, _warm)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)


# ── Cache ────────────────────────────────────────────────────────────────────

_report_cache: TTLCache = TTLCache(maxsize=64, ttl=300)  # 5 min TTL


def resolve_patient_id(patient_id: str) -> str:
    """Resolve raw patient/device IDs (like 934297-0122 or 934297-0134) to friendly display names."""
    if patient_id == "934297-0122":
        return "S (Chair)"
    if patient_id == "934297-0134":
        return "S (Bed)"
    return patient_id


# ── Client libraries (Round 28) ──────────────────────────────────────────────
# Patients are scoped by client/library, discovered from data/ sub-folders.
# A patient is loadable only under its own client (privacy boundary, R28_007).
# See backend/client_registry.py.

def _resolve_client(req_client: str | None, patient_id: str) -> str:
    """Pick the client a report runs under.

    If the request names a client, honour it verbatim (the privacy boundary):
    the patient must then exist under THAT client or the lookup 404s. Only when
    no client is named do we fall back to discovering the patient's owning
    client.
    """
    if req_client:
        if req_client not in discover_clients():
            raise HTTPException(status_code=404, detail=f"Unknown library '{req_client}'.")
        return req_client
    found = resolve_client_for_patient(patient_id)
    if found is None:
        raise HTTPException(status_code=404, detail=f"Patient '{patient_id}' not found in any library.")
    return found


def _cache_key(req: ReportRequest) -> str:
    raw = f"{req.client}|{req.patient_id}|{req.range_type}|{req.start}|{req.end}|{req.month}|{req.use_ai}"
    return hashlib.md5(raw.encode()).hexdigest()


# ── Pipeline ─────────────────────────────────────────────────────────────────

async def _run_pipeline(req: ReportRequest) -> tuple[dict, "pd.DataFrame"]:
    """Run the analysis pipeline.

    Steps:
      1. Look up patient
      2. Apply time window
      3. Run quality gates (REJECT → 422)
      4. Compute stats + data quality
      5. Detect episodes
      6. Compute triage + trend + action posture (BEFORE AI)
      7. Detect phases
      8. Compute report priority
      9. Generate narrative
     10. Generate charts
     11. Build response
    """
    import pandas as pd
    from .quality_gates import run_quality_gates
    from .window_intelligence import detect_phases, compute_report_priority
    from .models import Phase

    # ── Step 1: Look up patient (client-scoped) ──────────────────────────
    req.patient_id = resolve_patient_id(req.patient_id)
    client = _resolve_client(req.client, req.patient_id)
    _is_mh = client_is_csv(client)   # clean-CSV (MedHab-shape) client → month flow
    _mh_partial = False
    # Privacy boundary: load ONLY the selected client's folder. A patient that
    # belongs to another library is simply not present here → 404.
    all_data = load_client_data(client)
    if req.patient_id not in all_data:
        raise HTTPException(status_code=404,
            detail=f"Patient '{req.patient_id}' not found in library '{client}'.")

    df = all_data[req.patient_id]

    # ── Step 2: Apply time window ────────────────────────────────────────
    from .window_intelligence import find_most_interesting_week
    if _is_mh and req.range_type == "month":
        # MedHab month report — use the same per-patient-month window as the
        # batch (data-driven start/end, partial-period flag).
        spec = next((s for s in client_specs(client)
                     if s["patient"] == req.patient_id and s["month_key"] == req.month), None)
        if spec is None:
            raise HTTPException(status_code=404,
                detail=f"No data for {req.patient_id} in {req.month}.")
        df = apply_window(df, "custom", spec["start"], spec["end"])
        _mh_partial = spec["is_partial"]
    elif req.range_type == "smart_week":
        result = find_most_interesting_week(df)
        if result:
            ws_ts, we_ts = result["start"], result["end"]
            df = df[(df["timestamp"] >= ws_ts) & (df["timestamp"] <= we_ts)]
        else:
            # Fallback if no episodes found at all
            df = apply_window(df, "last_7d")
    else:
        # Non-month request (or a PAM patient): fall back to a standard window.
        rt = req.range_type if req.range_type != "month" else "last_1m"
        df = apply_window(df, rt, req.start, req.end)
    
    if df.empty:
        raise HTTPException(status_code=400, detail="No data in the selected time window.")

    # ── Step 1b: Prior Week Comparison (Now relative to window start) ────
    window_start_ts = df["timestamp"].min()
    window_end_ts = df["timestamp"].max()
    
    prior_start = window_start_ts - pd.Timedelta(days=7)
    prior_end = window_start_ts - pd.Timedelta(seconds=1)
    
    p_df = all_data[req.patient_id]
    p_df = p_df[(p_df["timestamp"] >= prior_start) & (p_df["timestamp"] <= prior_end)]
    
    prior_comparison = None
    if not p_df.empty:
        p_hr, p_rr = compute_stats(p_df)
        p_eps = detect_episodes(p_df)
        from .models import PriorComparison
        prior_comparison = PriorComparison(
            hr_avg=round(p_hr.mean, 1),
            rr_avg=round(p_rr.mean, 1),
            episode_count=len(p_eps),
            start_date=prior_start.strftime("%Y-%m-%d"),
            end_date=prior_end.strftime("%Y-%m-%d")
        )

    # ── Step 4: Run Quality Gates ────────────────────────────────────────
    # MedHab partial months render with a low-coverage badge rather than being
    # rejected (the min-substantive-days floor still applies).
    gate_result = run_quality_gates(df, window_start_ts, window_end_ts,
                                    downgrade_coverage_reject=_is_mh)

    if not gate_result["can_generate"]:
        raise HTTPException(
            status_code=422,
            detail=gate_result["reason"],
        )
    quality_warnings = gate_result["warnings"]

    # ── Step 5: Compute stats + daily aggregates ─────────────────────────
    hr_stats, rr_stats = compute_stats(df)
    full_stats = compute_full_stats(df)
    data_quality = compute_data_quality(df)
    data_resolution = compute_data_resolution(df)

    # ── Step 6: Detect episodes ──────────────────────────────────────────
    episodes = detect_episodes(df)

    # Rollups
    rollups = compute_rollups(episodes, df)

    # ── Step 7: Compute triage + trend + action posture ──────────────────
    # SAFETY BOUNDARY — computed before AI, cannot be changed by AI
    triage = compute_triage(episodes, rollups.coupled_fraction, df=df)
    trend, _ = compute_trend_assessment(df, episodes)

    # Max severity band and score
    max_band = "S0"
    max_severity_score = 0
    for ep in episodes:
        if ep.severity_score > max_severity_score:
            max_severity_score = ep.severity_score
        if ep.severity_band > max_band:
            max_band = ep.severity_band

    action_posture = compute_action_posture(triage, trend, rollups.coupled_fraction, max_band)

    # Window bounds
    window_start = df["timestamp"].min().strftime("%Y-%m-%d")
    window_end = df["timestamp"].max().strftime("%Y-%m-%d")

    # ── Step 8: Detect phases ────────────────────────────────────────────
    raw_phases = detect_phases(df, episodes)
    phases = [Phase(**p) for p in raw_phases]

    # ── Step 9: Compute report priority ──────────────────────────────────
    report_priority = compute_report_priority(
        episodes, raw_phases, max_severity_score, quality_warnings
    )
    # FIX 5 — keep the priority badge consistent with the triage tier (no
    # GREEN + HIGH contradiction). SKIP (data-quality) is preserved.
    if report_priority != "SKIP":
        from .config import RENDER_CONFIG as _RC
        report_priority = _RC.get("priority_by_triage", {}).get(triage, report_priority)

    # ── Step 9b: Detect sensor type & load bed data ──────────────────────
    sensor_type = "chair"
    bed_summary_model = None
    bed_summary_df = None
    alerts_df = None
    chart_bed_hours_b64 = ""

    if "location" in df.columns:
        locations = df["location"].unique()
        if Locations.BED in locations:
            sensor_type = "bed"

    # Also detect from bed excel presence — but only if it belongs to THIS patient
    raw_bed_df = load_bed_summary()
    raw_alerts_df = load_low_hr_alerts()
    if raw_bed_df is not None:
        # Verify bed file belongs to this patient by checking sheet names
        bed_file = _find_bed_excel()
        if bed_file:
            import openpyxl
            try:
                wb = openpyxl.load_workbook(str(bed_file), read_only=True)
                first_sheet = wb.sheetnames[0] if wb.sheetnames else ""
                wb.close()
                if req.patient_id in first_sheet:
                    sensor_type = "bed"
                # else: bed file exists but belongs to a different patient
            except Exception:
                pass

    if sensor_type == "bed" and raw_bed_df is not None:
        bed_summary_df = raw_bed_df.copy()
        alerts_df = raw_alerts_df

        # Filter bed summary to the report window
        ws_ts = pd.Timestamp(window_start)
        we_ts = pd.Timestamp(window_end)
        bed_summary_df = bed_summary_df[
            (bed_summary_df["date"] >= ws_ts) & (bed_summary_df["date"] <= we_ts)
        ].reset_index(drop=True)

        if alerts_df is not None and not alerts_df.empty:
            # Strip timezone info to match naive timestamps
            if alerts_df["timestamp"].dt.tz is not None:
                alerts_df = alerts_df.copy()
                alerts_df["timestamp"] = alerts_df["timestamp"].dt.tz_localize(None)
            alerts_df = alerts_df[
                (alerts_df["timestamp"] >= ws_ts) &
                (alerts_df["timestamp"] <= we_ts + pd.Timedelta(days=1))
            ].reset_index(drop=True)

        if not bed_summary_df.empty:
            # Build BedActivitySummary
            hours = bed_summary_df["hours_in_bed"].dropna()
            hr_lows = bed_summary_df["hr_low"].dropna()

            # Days above 16h
            days_above_16 = int((hours > 16).sum())

            # Alert stats
            total_alerts = len(alerts_df) if alerts_df is not None else 0
            alert_dates = set()
            if alerts_df is not None and not alerts_df.empty:
                alert_dates = set(alerts_df["timestamp"].dt.normalize().unique())

            # HR min on high-bed-days vs normal
            high_bed_mask = bed_summary_df["hours_in_bed"] > 16
            hr_min_high = bed_summary_df.loc[high_bed_mask, "hr_low"].dropna()
            hr_min_normal = bed_summary_df.loc[~high_bed_mask, "hr_low"].dropna()

            daily_data = []
            for _, row in bed_summary_df.iterrows():
                h = row["hours_in_bed"]
                if pd.isna(h):
                    color = "gray"
                elif h > 16:
                    color = "red"
                elif h >= 13:
                    color = "amber"
                else:
                    color = "green"

                daily_data.append(BedDaySummary(
                    date=row["date"].strftime("%Y-%m-%d"),
                    hours_in_bed=float(h) if not pd.isna(h) else 0,
                    hr_min=float(row["hr_low"]) if pd.notna(row.get("hr_low")) else 0,
                    has_alert=row["date"] in alert_dates,
                    color=color,
                ))

            bed_summary_model = BedActivitySummary(
                mean_daily_hours=float(hours.mean()) if len(hours) > 0 else 0,
                min_hours=float(hours.min()) if len(hours) > 0 else 0,
                max_hours=float(hours.max()) if len(hours) > 0 else 0,
                days_above_16h=days_above_16,
                alert_days=len(alert_dates),
                total_alerts=total_alerts,
                hr_min_high_bed_days=float(hr_min_high.mean()) if len(hr_min_high) > 0 else 0,
                hr_min_normal_days=float(hr_min_normal.mean()) if len(hr_min_normal) > 0 else 0,
                daily_data=daily_data,
            )

            # Generate bed hours chart
            chart_bed_hours_b64 = generate_bed_hours_chart(bed_summary_df, alerts_df)

    # ── Step 9c: Activity & Location Stats ───────────────────────────────
    positional_stats = compute_positional_stats(df)
    activity_data = compute_activity_data(df)

    # ── Step 10: Generate narrative ──────────────────────────────────────
    narrative, actions, narrative_source = await generate_narrative(
        req.patient_id, window_start, window_end,
        hr_stats, rr_stats, data_quality,
        episodes, rollups, triage, trend, action_posture,
        use_llm_override=req.use_ai,
        quality_warnings=quality_warnings,
        phases=raw_phases,
        bed_summary=bed_summary_model,
        activity_trend=activity_data,
        positional_stats=positional_stats,
    )

    # Cap actions at max_actions
    actions = actions[:settings.max_actions]

    # ── Step 11: Generate charts ─────────────────────────────────────────
    chart_b64 = generate_combined_chart(df, episodes)
    histogram_b64 = generate_histogram(df, hr_p5=hr_stats.p5, hr_p95=hr_stats.p95)

    chart_positional_b64 = generate_positional_chart(df)
    chart_activity_b64 = generate_activity_trend_chart(df)

    # ── Coverage Summary ─────────────────────────────────────────────────
    # Build coverage summary — never show > 100%.
    # For multi-sensor patients the combined row count exceeds calendar hours,
    # so report per-sensor instead (each sensor vs. the same expected hours).
    if positional_stats and len(positional_stats.rows) >= 1:
        expected_h = data_quality.expected_hours
        parts = []
        for row in positional_stats.rows:
            loc_hours = row.hours
            loc_pct = min(round(loc_hours / max(expected_h, 1) * 100, 1), 100.0)
            parts.append(f"{row.location}: {loc_hours}/{expected_h}h ({loc_pct}%)")
        coverage_summary = "  |  ".join(parts)
    else:
        capped_pct = min(data_quality.quality_pct, 100.0)
        coverage_summary = f"{data_quality.total_hours}/{data_quality.expected_hours}h ({capped_pct}%)"

    # ── Last-24h triage layer (Round 28) ─────────────────────────────────
    # The snapshot stats plus the 24h episodic events, one-line summary, and
    # 24h status banner classification — same engines, scoped to the last 24h.
    snapshot_24h = compute_last_24h_snapshot(df)
    layer_24h = await compute_24h_layer(req.patient_id, df)
    if snapshot_24h is not None and layer_24h is not None:
        snapshot_24h.update(layer_24h)
    # The 30-day tier travels alongside so the banner can label each window and
    # the two never read as a contradiction (header tier vs 24h banner).
    if snapshot_24h is not None:
        snapshot_24h["status_30d"] = triage

    print(f"DEBUG PIPELINE: hr_stats is {hr_stats}")
    report_dict = {
        "patient_id": req.patient_id,
        "window_start": window_start,
        "window_end": window_end,
        "report_date": (window_end_ts + pd.Timedelta(days=1)).strftime("%Y-%m-%d"),
        "data_resolution": data_resolution,
        "coverage_summary": coverage_summary,
        "snapshot_24h": snapshot_24h,
        "disclaimer": "Measurement data only. Not a diagnosis. Values reflect selected windows where heart rate or breathing was outside the stated ranges, from radar based vital sign measurement.",
        "hr_summaries": hr_stats.model_dump() if hasattr(hr_stats, "model_dump") else hr_stats,
        "rr_summaries": rr_stats.model_dump() if hasattr(rr_stats, "model_dump") else rr_stats,
        "full_stats": full_stats.model_dump() if full_stats else None,
        "data_quality": data_quality.model_dump() if hasattr(data_quality, "model_dump") else data_quality,
        "episodes": [e.model_dump() if hasattr(e, "model_dump") else e for e in episodes[:settings.max_events_table]],
        "episode_rollups": rollups.model_dump() if hasattr(rollups, "model_dump") else rollups,
        "triage": triage,
        "trend_assessment": trend,
        "overall_action_posture": action_posture,
        "max_severity_score": max_severity_score,
        "narrative": narrative,
        "suggested_actions": actions,
        "use_ai": req.use_ai,
        "narrative_source": narrative_source,
        "report_priority": report_priority,
        "phases": [p.model_dump() if hasattr(p, "model_dump") else p for p in phases],
        "quality_warnings": quality_warnings,
        "positional_comparison": positional_stats.model_dump() if positional_stats else None,
        "activity_trend": activity_data.model_dump() if activity_data else None,
        "chart_combined_b64": chart_b64,
        "chart_histogram_b64": histogram_b64,
        "chart_positional_b64": chart_positional_b64,
        "chart_activity_b64": chart_activity_b64,
        "sensor_type": sensor_type,
        "bed_summary": bed_summary_model.model_dump() if bed_summary_model else None,
        "chart_bed_hours_b64": chart_bed_hours_b64,
        "prior_comparison": prior_comparison.model_dump() if prior_comparison else None,
        # MedHab cohort: 30-day report label, partial-period note, end-of-period
        # clustering — mirrors the batch so the live preview matches the PDF.
        "report_label": "30DayPeriod" if _is_mh else "",
        "is_fallback_90d": _mh_partial,
        "end_of_period": compute_end_of_period_clustering(
            episodes, window_start_ts, window_end_ts),
    }
    return report_dict, df


# ── Endpoints ────────────────────────────────────────────────────────────────

@app.get("/api/clients")
async def list_clients():
    """List the patient libraries (clients), discovered from data/ sub-folders.

    Each library is a separate client; selecting one scopes the patient list so
    one client's data can never appear under another (Round 28, Item 3).
    """
    try:
        return {"clients": [{"id": c, "label": client_label(c)} for c in discover_clients()]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/summary")
async def library_summary(client: str | None = None):
    """24h library summary — every patient in one library, critical first, 24h
    only. Same engine as the per-patient banner (statuses match). Scoped to a
    single client; never mixes libraries."""
    try:
        from .library_summary import build_library_summary
        clients = discover_clients()
        if not clients:
            return {"client": None, "label": "", "patients": []}
        if client is None:
            client = clients[0]
        if client not in clients:
            raise HTTPException(status_code=404, detail=f"Unknown library '{client}'.")
        return build_library_summary(client)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/summary/pdf")
async def library_summary_pdf(client: str | None = None):
    """Downloadable PDF of a library's 24h summary (same renderer as the batch)."""
    try:
        from datetime import datetime
        from .library_summary import build_library_summary
        from .pdf_render import generate_library_summary_pdf
        clients = discover_clients()
        if not clients:
            raise HTTPException(status_code=404, detail="No libraries found.")
        if client is None:
            client = clients[0]
        if client not in clients:
            raise HTTPException(status_code=404, detail=f"Unknown library '{client}'.")
        summary = build_library_summary(client)
        pdf = generate_library_summary_pdf(summary, datetime.now().strftime("%B %d, %Y"))
        fname = f"CardioReport_{client}_24h_Summary.pdf"
        return Response(content=pdf, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="{fname}"'})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/patients")
async def list_patients_endpoint(client: str | None = None):
    """Return the patient list for a library. ``?client=<id>`` scopes the list
    to exactly that client; omitting it defaults to the first discovered
    library. A patient from one library never appears under another."""
    try:
        clients = discover_clients()
        if not clients:
            return {"client": None, "patients": []}
        if client is None:
            client = clients[0]
        if client not in clients:
            raise HTTPException(status_code=404, detail=f"Unknown library '{client}'.")
        return {"client": client, "patients": list_patients(client)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/months")
async def list_months(client: str | None = None):
    """Available report months for a (CSV-shape) library, discovered from the
    data. Excel-shape libraries (e.g. PAM Health) report over rolling ranges,
    not months, so this returns an empty list and the UI shows range options."""
    try:
        clients = discover_clients()
        if not clients:
            return {"months": []}
        if client is None:
            client = clients[0]
        if client not in clients:
            raise HTTPException(status_code=404, detail=f"Unknown library '{client}'.")
        seen: dict[str, str] = {}
        for s in client_specs(client):
            seen[s["month_key"]] = s["month_label"].replace("_", " ")
        return {"months": [{"key": k, "label": seen[k]} for k in sorted(seen)]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/patients/{patient_id}/locations")
async def patient_locations(patient_id: str, client: str | None = None):
    """Return the locations (Chair, Bed, Living Room) and date range for a patient.

    Enables the frontend to intelligently disable unavailable report types.
    """
    try:
        patient_id = resolve_patient_id(patient_id)
        client = _resolve_client(client, patient_id)
        if client_is_csv(client):
            # CSV-shape client (MedHab) — a single wearable-style stream, no
            # positional sensors.
            df = load_client_data(client)[patient_id]
            ts = df["timestamp"]
            return {
                "patient_id": patient_id,
                "locations": [],
                "sensor_types": [],
                "date_range": {"start": ts.min().strftime("%Y-%m-%d"),
                               "end": ts.max().strftime("%Y-%m-%d")},
                "total_hours": int(len(df)),
            }
        meta = get_patient_metadata(patient_id)
        if not meta["locations"]:
            raise HTTPException(status_code=404, detail=f"Patient '{patient_id}' not found or has no data.")
        return meta
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/patients/{patient_id}/interesting-week")
async def interesting_week(patient_id: str, client: str | None = None):
    """Find the 7-day window with the highest clinical burden for a patient.

    Slides a 7-day window across the entire dataset, runs episode detection
    on each position, and returns the top-scoring result.
    """
    from .window_intelligence import find_most_interesting_week

    patient_id = resolve_patient_id(patient_id)
    client = _resolve_client(client, patient_id)
    all_data = load_client_data(client)
    if patient_id not in all_data:
        raise HTTPException(status_code=404, detail=f"Patient '{patient_id}' not found.")

    df = all_data[patient_id]
    result = find_most_interesting_week(df)

    if result is None:
        raise HTTPException(status_code=404, detail="No window with sufficient data found.")

    return result

@app.post("/api/report/preview")
async def report_preview(req: ReportRequest):
    """Generate and return the report as JSON for web preview."""
    req.patient_id = resolve_patient_id(req.patient_id)
    key = _cache_key(req)
    if key in _report_cache:
        return _report_cache[key]

    report_dict, df = await _run_pipeline(req)
    from fastapi.encoders import jsonable_encoder
    result = jsonable_encoder(report_dict)
    _report_cache[key] = result
    return result


@app.post("/api/report/pdf")
async def report_pdf(req: ReportRequest):
    """Generate and return the PDF report."""
    req.patient_id = resolve_patient_id(req.patient_id)
    from .pdf_render import _v
    report_obj, df = await _run_pipeline(req)

    episodes = _v(report_obj, "episodes", [])
    pdf_bytes = generate_pdf(report_obj, df=df, episodes=episodes)

    pid = _v(report_obj, "patient_id")
    ws = _v(report_obj, "window_start")
    # Filename: CardioReport_<patient>_<Month_Year> (easy for users).
    try:
        import pandas as _pd
        month_label = _pd.Timestamp(ws).strftime("%B_%Y") if ws else ""
    except Exception:
        month_label = ""
    filename = f"CardioReport_{pid}_{month_label}.pdf" if month_label else f"CardioReport_{pid}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/report/events.json")
async def export_events(patient_id: str, range_type: Literal["last_24h", "last_7d", "last_15d", "last_1m", "last_3m", "custom", "smart_week"] = "last_3m",                        start: Optional[str] = None, end: Optional[str] = None):
    """Export detected episodes as JSON."""
    patient_id = resolve_patient_id(patient_id)
    req = ReportRequest(patient_id=patient_id, range_type=range_type,
                        start=start, end=end)
    report, _ = await _run_pipeline(req)
    return {
        "patient_id": report["patient_id"],
        "window": {"start": report["window_start"], "end": report["window_end"]},
        "triage": report["triage"],
        "trend_assessment": report["trend_assessment"],
        "episodes": report["episodes"],
        "rollups": report["episode_rollups"],
    }


@app.get("/api/health")
async def health():
    return {"status": "ok", "version": settings.app_version}
